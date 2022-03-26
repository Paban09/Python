from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter






#wb=load_workbook('tim.xlsx')
#ws=wb.active
#ws.move_range("C1:D11", rows=2, cols=2)
#wb.save('tim.xlsx')






#wb=load_workbook('Grades.xlsx')
#ws=wb.active
#ws.delete_rows(7)
#wb.save('Grades.xlsx')





#wb=load_workbook('tim.xlsx')
#ws=wb.active
#ws.unmerge_cells("A1:D1") #to merge the cells 
#ws.unmerge_cells("A1:D1") #to unmerge the cells 

#wb.save('tim.xlsx')






#wb=load_workbook('tim.xlsx')
#ws=wb.active
#for row in range(1,11):
#    for col in range(1,5):
#        char = get_column_letter(col)
#        print(ws[char+str(row)].value)    






#wb=Workbook()
#ws=wb.active #default workshhet while creating a workbook
#ws.title="Data"

#ws.append(['Tim','Is','Great','!'])
#ws.append(['Tim','Is','Great','!'])
#ws.append(['Tim','Is','Great','!'])
#ws.append(['Tim','Is','Great','!'])
#ws.append(['end'])

#wb.save('tim.xlsx')






#wb=load_workbook('Grades.xlsx') #wb defines as workbook
#ws=wb.active #ws defines as worksheet
#print(ws) #how many active sheet is availabe in xl file
#print(ws['A1'].value)  #access a value
#ws['A2'].value="Test" #Change a value
#wb.save('Grades.xlsx') #saving the shhet to look upto the changes
#print(wb.sheetnames)
